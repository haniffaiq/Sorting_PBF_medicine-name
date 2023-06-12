import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import math

# Baca data dari file CSV
df = pd.read_excel('guide supplier kwb.xlsx')
df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
cols_to_check = ['Nama Barang', 'Satuan', 'Supplier Termurah', 'HPP Min', 'HPP Max', 'HPP Average', 'Harga Jual Min', 'Harga Jual Max']
df = df.dropna(subset=cols_to_check, how='all')
df['Nama Barang'] = df['Nama Barang'].str.lower()
df = df.sort_values(by='Supplier Termurah', key=lambda x: x.str[0])
df = df.fillna("PBF Tidak Ada")

print(df)




def find_matching_product(product_name):
    matching_rows = df[df['Nama Barang'].str.contains(product_name, case=False, regex=False, na=False)]
    if not matching_rows.empty:
        tempat_order_list = []
        tempat_order_1 = matching_rows['Supplier Termurah'].dropna().apply(lambda x: x.strip()).values.tolist()
        tempat_order_list.extend(tempat_order_1)
        return tempat_order_list
    else:
        return None
    
# Hapus baris dengan nilai NaN pada kolom yang dibutuhkan
df_lapstok = pd.read_excel('LapStokDibawahLimit 07062023.xls', header=6)
df_lapstok = df_lapstok.drop(df_lapstok.columns[df_lapstok.columns.str.contains('unnamed', case=False)], axis=1)
cols_to_check_lapstok = ['PLU', 'Nama Produk', 'Satuan', 'Qty', 'Stok Min', 'Stok Max', 'Pesan']
df_lapstok = df_lapstok.dropna(subset=cols_to_check_lapstok)

# Ubah kolom 'Nama Produk' menjadi lowercase
df_lapstok['Nama Produk'] = df_lapstok['Nama Produk'].str.lower()

# Inisialisasi list untuk menampung hasil
result_rows = []
for index, row in tqdm(df_lapstok.iterrows(), total=len(df_lapstok), desc='Processing'):
    nama_produk = row['Nama Produk']
    pesan = row['Pesan']

    # Cari tempat order menggunakan fungsi find_matching_product
    tempat_order = find_matching_product(nama_produk)

    # Jika ada tempat order yang cocok
    if tempat_order:
        jumlah_pesan = pesan  # Gunakan nilai pesan dari file LapStok.xlsx

        # Tambahkan baris ke list hasil
        result_rows.append({
            'Nama Produk': nama_produk,
            'Tempat Order': tempat_order,
            'Jumlah Pesan': jumlah_pesan
        })
df_tempat_order = pd.DataFrame(result_rows, columns=['Nama Produk', 'Tempat Order', 'Jumlah Pesan'])
print(df_tempat_order)

df_tempat_order['Tempat Order'] = df_tempat_order['Tempat Order'].astype(str)


df_grouped = df_tempat_order.groupby('Tempat Order').apply(lambda x: x.apply(lambda y: pd.Series(y[y.notnull()]), axis=1)).reset_index(drop=True)

print(df_grouped)

def export_satu():
    # Membuat workbook dan worksheet baru
    wb = Workbook()
    ws = wb.active

    # Menulis dataframe ke dalam worksheet
    for row in dataframe_to_rows(df_tempat_order, index=False, header=True):
        ws.append(row)

    # Mengatur lebar kolom
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Menyimpan workbook ke dalam file Excel
    wb.save('Hasil.xlsx')

    print("DataFrame exported to Excel successfully!")

def export_dua():
    # Membuat workbook dan worksheet baru
    wb = Workbook()
    ws = wb.active

    # Menulis dataframe ke dalam worksheet
    for row in dataframe_to_rows(df_grouped, index=False, header=True):
        ws.append(row)

    # Mengatur lebar kolom
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Menyimpan workbook ke dalam file Excel
    wb.save('Hasil2.xlsx')

    print("DataFrame exported to Excel successfully!")



obatTidakDijual = pd.read_excel('DataObatTidakDijual KWB.xlsx')
obatTidakDijual = obatTidakDijual.fillna(0)

# Buat dataframe baru dengan data yang Dijual = '1'
data_tidak_dijual = obatTidakDijual.loc[obatTidakDijual['Dijual'] == 1].copy()
data_tidak_dijual = data_tidak_dijual.drop(columns=data_tidak_dijual.columns[data_tidak_dijual.columns.str.contains('Unnamed')])

data_tidak_dijual['Nama Produk'] = data_tidak_dijual['Nama Produk'].str.lower()

df_grouped = df_grouped[~df_grouped['Nama Produk'].isin(data_tidak_dijual['Nama Produk'])]

# Cetak dataframe baru
print(data_tidak_dijual)
export_satu()
export_dua()