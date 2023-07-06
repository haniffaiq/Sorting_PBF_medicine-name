import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import datetime
import locale


def format_tanggal():
    # Mengatur lokasi menjadi Indonesia
    locale.setlocale(locale.LC_ALL, 'id_ID')
    
    tanggal_sekarang = datetime.datetime.now()
    format_tanggal = tanggal_sekarang.strftime("%d-%B-%Y")
    
    return format_tanggal


# Fungsi untuk mencari produk yang cocok
def find_matching_product(df_supplier, product_name):
    matching_rows = df_supplier[df_supplier['Nama Barang'].str.contains(product_name, case=False, regex=False, na=False)]
    if not matching_rows.empty:
        tempat_order_list = matching_rows['Supplier Termurah'].dropna().str.strip().tolist()
        return tempat_order_list
    else:
        return None

# Membaca data supplier
def read_supplier_data(file_path):
    df_supplier = pd.read_excel(file_path)
    df_supplier = df_supplier.loc[:, ~df_supplier.columns.str.contains('^Unnamed')]
    cols_to_check = ['Nama Barang', 'Satuan', 'Supplier Termurah', 'HPP Min', 'HPP Max', 'HPP Average', 'Harga Jual Min', 'Harga Jual Max']
    df_supplier = df_supplier.dropna(subset=cols_to_check, how='all')
    df_supplier['Nama Barang'] = df_supplier['Nama Barang'].str.lower()
    df_supplier = df_supplier.sort_values(by='Supplier Termurah', key=lambda x: x.str[0])
    df_supplier = df_supplier.fillna("PBF Tidak Ada")
    return df_supplier

# Membaca data lapstok
def read_lapstok_data(file_path):
    df_lapstok = pd.read_excel(file_path, header=6)
    df_lapstok = df_lapstok.drop(df_lapstok.columns[df_lapstok.columns.str.contains('unnamed', case=False)], axis=1)
    cols_to_check_lapstok = ['PLU', 'Nama Produk', 'Satuan', 'Qty', 'Stok Min', 'Stok Max', 'Pesan']
    df_lapstok = df_lapstok.dropna(subset=cols_to_check_lapstok)
    df_lapstok['Nama Produk'] = df_lapstok['Nama Produk'].str.lower()
    return df_lapstok

# Membaca data obat tidak dijual
def read_obat_tidak_dijual_data(file_path):
    df_obat_tidak_dijual = pd.read_excel(file_path)
    df_obat_tidak_dijual = df_obat_tidak_dijual.fillna(0)
    df_obat_tidak_dijual = df_obat_tidak_dijual.loc[df_obat_tidak_dijual['Dijual'] == 1].copy()
    df_obat_tidak_dijual = df_obat_tidak_dijual.drop(columns=df_obat_tidak_dijual.columns[df_obat_tidak_dijual.columns.str.contains('Unnamed')])
    df_obat_tidak_dijual['Nama Produk'] = df_obat_tidak_dijual['Nama Produk'].str.lower()
    return df_obat_tidak_dijual

# Menggabungkan data dan melakukan pemrosesan
def process_data(supplier_file, lapstok_file, obat_tidak_dijual_file,  hasil_file, hasil2_file):
    # Baca data dari file supplier
    df_supplier = read_supplier_data(supplier_file)

    # Baca data dari file lapstok
    df_lapstok = read_lapstok_data(lapstok_file)

    # Baca data dari file obat tidak dijual
    df_obat_tidak_dijual = read_obat_tidak_dijual_data(obat_tidak_dijual_file)

    # Inisialisasi list untuk menampung hasil
    result_rows = []
    for index, row in df_lapstok.iterrows():
        nama_produk = row['Nama Produk']
        pesan = row['Pesan']

        # Cari tempat order menggunakan fungsi find_matching_product
        tempat_order = find_matching_product(df_supplier, nama_produk)

        # Jika ada tempat order yang cocok
        if tempat_order:
            jumlah_pesan = pesan

            # Tambahkan baris ke list hasil
            result_rows.append({
                'Nama Produk': nama_produk,
                'Tempat Order': tempat_order,
                'Jumlah Pesan': jumlah_pesan
            })

    # Buat DataFrame hasil
    df_tempat_order = pd.DataFrame(result_rows, columns=['Nama Produk', 'Tempat Order', 'Jumlah Pesan'])

    # Hapus produk yang tidak dijual
    df_tempat_order = df_tempat_order[~df_tempat_order['Nama Produk'].isin(df_obat_tidak_dijual['Nama Produk'])]
    df_tempat_order['Tempat Order'] = df_tempat_order['Tempat Order'].astype(str)


    df_grouped = df_tempat_order.groupby('Tempat Order').apply(lambda x: x.apply(lambda y: pd.Series(y[y.notnull()]), axis=1)).reset_index(drop=True)

    # print(df_grouped)
    # # Export DataFrame ke file Excel

    # export_to_excel(df_tempat_order, hasil_file)
    export_to_excel(df_grouped, hasil2_file)

    # # Cetak DataFrame obat tidak dijual
    # print(df_obat_tidak_dijual)

# Fungsi untuk mengekspor DataFrame ke file Excel
def export_to_excel(df, file_name):
    wb = Workbook()
    ws = wb.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    wb.save(file_name)
    print("DataFrame exported to Excel successfully!")






pathDatabase_KWM = './Database PBF/Database PBF KWM.xlsx'
pathDataObatTidakDijual_KWM = './Data Obat Tidak Dijual/DataObatTidakDijual KWM.xlsx'
pathLapStok_KWM = './LapStok/Lapstok KWM.xls'

pathDatabase_KWB = './Database PBF/Database PBF KWB.xlsx'
pathDataObatTidakDijual_KWB = './Data Obat Tidak Dijual/DataObatTidakDijual KWB.xlsx'
pathLapStok_KWB = './LapStok/Lapstok KWB.xls'

pathDatabase_GEDA = './Database PBF/Database PBF GEDA.xlsx'
pathDataObatTidakDijual_GEDA = './Data Obat Tidak Dijual/DataObatTidakDijual GEDA.xlsx'
pathLapStok_GEDA = './LapStok/Lapstok GEDA.xls'

pathDatabase_MUT = './Database PBF/Database PBF MUT.xlsx'
pathDataObatTidakDijual_MUT = './Data Obat Tidak Dijual/DataObatTidakDijual MUT.xlsx'
pathLapStok_MUT = './LapStok/Lapstok MUT.xls'

pathDatabase_MUS = './Database PBF/Database PBF MUS.xlsx'
pathDataObatTidakDijual_MUS = './Data Obat Tidak Dijual/DataObatTidakDijual MUS.xlsx'
pathLapStok_MUS = './LapStok/Lapstok MUS.xls'

pathDatabase_KWU = './Database PBF/Database PBF KWU.xlsx'
pathDataObatTidakDijual_KWU = './Data Obat Tidak Dijual/DataObatTidakDijual KWU.xlsx'
pathLapStok_KWU = './LapStok/Lapstok KWU.xls'



# Hari ini
tanggal_hari_ini = format_tanggal()
print(tanggal_hari_ini)


def cek_file(folder, file):
    # Mengecek apakah file ada di folder
    file_path = os.path.join(folder, file)
    if (os.path.isfile(file_path)):
        if(file == "Lapstok KWM.xls"):
            print("===================================================================")
            print("File {}  ditemukan di folder LapStok".format(file))
            process_data(pathDatabase_KWM, pathLapStok_KWM, pathDataObatTidakDijual_KWM, "./Result/HASIL_KWM_1.xlsx", "./Result/Order-KWM-{}.xlsx".format(tanggal_hari_ini))
            print("File diexport ./Result/Order-KWM-{}.xlsx".format(tanggal_hari_ini))
            print("===================================================================")
        elif(file == "Lapstok KWB.xls"):
            print("===================================================================")
            print("File {}  ditemukan di folder LapStok".format(file))
            process_data(pathDatabase_KWB, pathLapStok_KWB, pathDataObatTidakDijual_KWB, "./Result/HASIL_KWB_1.xlsx", "./Result/Order-KWB-{}.xlsx".format(tanggal_hari_ini))
            print("File diexport ./Result/Order-KWB-{}.xlsx".format(tanggal_hari_ini))
            print("===================================================================")
        elif(file == "Lapstok GEDA.xls"):
            print("===================================================================")
            print("File {}  ditemukan di folder LapStok".format(file))
            process_data(pathDatabase_GEDA, pathLapStok_GEDA, pathDataObatTidakDijual_GEDA, "./Result/HASIL_GEDA_1.xlsx", "./Result/Order-GEDA-{}.xlsx".format(tanggal_hari_ini))
            print("File diexport ./Result/Order-GEDA-{}.xlsx".format(tanggal_hari_ini))
            print("===================================================================")

        elif(file == "Lapstok MUT.xls"):
            print("===================================================================")
            print("File {}  ditemukan di folder LapStok".format(file))
            process_data(pathDatabase_MUT, pathLapStok_MUT, pathDataObatTidakDijual_MUT, "./Result/HASIL_MUT_1.xlsx", "./Result/Order-MUT-{}.xlsx".format(tanggal_hari_ini))
            print("File diexport ./Result/Order-MUT-{}.xlsx".format(tanggal_hari_ini))
            print("===================================================================")
        elif(file == "Lapstok MUS.xls"):
            print("===================================================================")
            print("File {}  ditemukan di folder LapStok".format(file))
            process_data(pathDatabase_MUS, pathLapStok_MUS, pathDataObatTidakDijual_MUS, "./Result/HASIL_MUS_1.xlsx", "./Result/Order-MUS-{}.xlsx".format(tanggal_hari_ini))
            print("File diexport ./Result/Order-MUS-{}.xlsx".format(tanggal_hari_ini))
            print("===================================================================")
        elif(file == "Lapstok KWU.xls"):
            print("===================================================================")
            print("File {}  ditemukan di folder LapStok".format(file))
            process_data(pathDatabase_KWU, pathLapStok_KWU, pathDataObatTidakDijual_KWU, "./Result/HASIL_KWU_1.xlsx", "./Result/Order-KWU-{}.xlsx".format(tanggal_hari_ini))
            print("File diexport ./Result/Order-KWU-{}.xlsx".format(tanggal_hari_ini))
            print("===================================================================")


    else:
        print("===================================================================")
        print("File {} tidak ditemukan di folder LapStok".format(file))
        print("===================================================================")

# Memasukkan nama folder dan nama file yang ingin dicek

folderLaptStok = "./Lapstok"
fileKWM = "Lapstok KWM.xls"
fileKWB = "Lapstok KWB.xls"
fileGEDA = "Lapstok GEDA.xls"
fileMUT = "Lapstok MUT.xls"
fileMUS = "Lapstok MUS.xls"
fileKWU = "Lapstok KWU.xls"

cek_file(folderLaptStok, fileKWM)
cek_file(folderLaptStok, fileKWB)
cek_file(folderLaptStok, fileGEDA)
cek_file(folderLaptStok, fileMUT)
cek_file(folderLaptStok, fileMUS)
cek_file(folderLaptStok, fileKWU)



# # Jalankan proses data
# # ====CODE DIATAS TULISAN INI JANGAN DIUBAH !!!!====


# # === PROSES KWM ===
# process_data(pathDatabase_KWM, pathLapStok_KWM, pathDataObatTidakDijual_KWM, "./Result/HASIL_KWM_1.xlsx", "./Result/Order-KWM-{}.xlsx".format(tanggal_hari_ini))

# # === PROSES KWB ===
# process_data(pathDatabase_KWB, pathLapStok_KWB, pathDataObatTidakDijual_KWB, "./Result/HASIL_KWB_1.xlsx", "./Result/Order-KWB-{}.xlsx".format(tanggal_hari_ini))

# # === PROSES GEDA ===
# process_data(pathDatabase_GEDA, pathLapStok_GEDA, pathDataObatTidakDijual_GEDA, "./Result/HASIL_GEDA_1.xlsx", "./Result/Order-GEDA-{}.xlsx".format(tanggal_hari_ini))
